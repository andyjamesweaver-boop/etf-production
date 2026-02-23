"""
ASX Investment Products Monthly Report scraper.

Downloads the Excel spreadsheet from ASX and parses the ETF sheets to
extract the master list of ASX-listed ETFs with FUM and flow data.
"""

import os
import re
import logging
from datetime import datetime

from scrapers.config import (
    ASX_INVESTMENT_PRODUCTS_INDEX_URL, ASX_INVESTMENT_PRODUCTS_BASE_URL, DATA_DIR,
    normalise_issuer, normalise_asset_class,
)
from scrapers.base_scraper import download_file, fetch
from scrapers.db_writer import get_connection, upsert_etfs, log_scrape, update_issuer_stats

logger = logging.getLogger(__name__)


def _find_latest_excel_url() -> str | None:
    """
    Scrape the ASX monthly report index page to find the most recent Excel file URL.
    Returns the full URL or None.
    """
    try:
        from bs4 import BeautifulSoup
    except ImportError:
        logger.error("beautifulsoup4 required to discover ASX report URL")
        return None

    resp = fetch(ASX_INVESTMENT_PRODUCTS_INDEX_URL)
    if not resp:
        logger.error("Could not fetch ASX report index page")
        return None

    soup = BeautifulSoup(resp.text, 'html.parser')
    for a in soup.find_all('a', href=True):
        href = a['href']
        if '.xlsx' in href and 'asx-investment-products' in href:
            # Return the first (most recent) Excel link
            if href.startswith('http'):
                return href
            return ASX_INVESTMENT_PRODUCTS_BASE_URL + href

    logger.error("No Excel link found on ASX report index page")
    return None


def _download_report() -> str | None:
    """Download the latest ASX monthly report Excel file. Returns path or None."""
    os.makedirs(DATA_DIR, exist_ok=True)
    ts = datetime.now().strftime('%Y%m')
    dest = os.path.join(DATA_DIR, f'asx_investment_products_{ts}.xlsx')

    # Re-use if downloaded this month
    if os.path.exists(dest):
        age_hours = (datetime.now().timestamp() - os.path.getmtime(dest)) / 3600
        if age_hours < 12:
            logger.info(f"Using cached report: {dest}")
            return dest

    url = _find_latest_excel_url()
    if not url:
        return None

    logger.info(f"Downloading ASX report from: {url}")
    if download_file(url, dest):
        return dest
    return None


def _safe_float(val) -> float | None:
    """Convert a cell value to float, or None."""
    if val is None:
        return None
    try:
        s = str(val).replace(',', '').replace('$', '').replace('%', '').strip()
        if s in ('', '-', 'N/A', 'n/a'):
            return None
        return float(s)
    except (ValueError, TypeError):
        return None


def _find_header_row(ws, markers: list[str]) -> int | None:
    """Find the row index containing header markers (case-insensitive)."""
    markers_lower = [m.lower() for m in markers]
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=30, values_only=False), start=1):
        cell_texts = [str(c.value or '').strip().lower() for c in row]
        if any(m in ' '.join(cell_texts) for m in markers_lower):
            return row_idx
    return None


def _from_decimal_pct(v) -> float | None:
    """
    Some ASX report columns store returns/yields as decimal fractions (0.085 = 8.5%).
    If the value looks like a fraction (abs < 2.0), multiply by 100.
    Values already in percent form (e.g. 8.5) are left unchanged.
    """
    if v is None:
        return None
    if -2.0 < v < 2.0:
        return round(v * 100, 4)
    return v


def _parse_etf_sheet(ws) -> list[dict]:
    """
    Parse a worksheet that lists ETFs.
    Handles both generic ETF sheets and the ASX 'Spotlight ETP List' format which
    carries fund flows, full return history, price, and distribution yield.
    """
    header_row = _find_header_row(ws, ['ticker', 'code', 'fund name', 'etf name'])
    if header_row is None:
        header_row = _find_header_row(ws, ['asx code', 'product name'])
    if header_row is None:
        logger.debug(f"No header row found in sheet '{ws.title}'")
        return []

    # Read header cells
    headers = []
    for cell in ws[header_row]:
        headers.append(str(cell.value or '').strip().lower())

    # Map header names to column indices.
    # Order matters: flow_1y is checked before flow_1m so "12M Funds Inflow" gets
    # claimed by flow_1y before the broader "inflow" pattern claims it for flow_1m.
    col_map = {}
    patterns = {
        'code':      re.compile(r'(code|ticker|asx\s*code)'),
        'name':      re.compile(r'(fund\s*name|product\s*name|etf\s*name|name)'),
        'issuer':    re.compile(r'(issuer|manager|provider|fund\s*manager)'),
        'category':  re.compile(r'(category|asset\s*class|sector|classification)'),
        'fum':       re.compile(r'(\bfum\b|fund\s*size|\baum\b|net\s*assets|total\s*assets)'),
        # Flows — match 12M variant first so the 1M pattern doesn't consume it
        'flow_1y':   re.compile(r'(12\s*m.*inflow|12.*month.*flow|1\s*y(ear)?\s*flow|annual\s*flow)'),
        'flow_1m':   re.compile(r'(funds?\s+inflow|inflow.*outflow|monthly\s*flow|1\s*m(onth)?\s*flow)'),
        'flow_3m':   re.compile(r'(3\s*m(onth)?\s*flow|quarterly\s*flow)'),
        # Returns — "1 Month Total Return", "1 Year Total Return", etc.
        'return_1m': re.compile(r'(1\s*m(onth)?\s*(total\s*)?return)'),
        'return_1y': re.compile(r'(1\s*y(ear)?\s*(total\s*)?return)'),
        'return_3y': re.compile(r'(3\s*y(ear)?\s*(total\s*)?return)'),
        'return_5y': re.compile(r'(5\s*y(ear)?\s*(total\s*)?return)'),
        # Price — "Last ($)" column
        'price':     re.compile(r'(\blast\s*\(\$\)|\blast\s*\$|\blast\s+price|\bclose\s+price)'),
        # Yield — "Historical Distribution Yield"
        'yield':     re.compile(r'(distribution\s*yield|historical.*yield|dividend.*yield)'),
        # MER — already in percent form in the spreadsheet (e.g. 0.04 means 0.04%)
        'mer':       re.compile(r'(\bmer\b|management\s*fee|expense\s*ratio)'),
    }

    for idx, h in enumerate(headers):
        for field, pat in patterns.items():
            if field not in col_map and pat.search(h):
                col_map[field] = idx

    if 'code' not in col_map:
        logger.debug(f"No 'code' column in sheet '{ws.title}'")
        return []

    results = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        code_val = row[col_map['code']] if col_map.get('code') is not None and col_map['code'] < len(row) else None
        if not code_val:
            continue
        code = str(code_val).strip().upper()
        if not re.match(r'^[A-Z0-9]{2,6}$', code) or code.isdigit():
            continue

        def cell(field):
            idx = col_map.get(field)
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        etf = {
            'code': code,
            'name': str(cell('name') or '').strip() or None,
            'issuer': normalise_issuer(str(cell('issuer') or '').strip() or None),
            'asset_class': normalise_asset_class(str(cell('category') or '').strip() or None),
            'exchange': 'ASX',
            'fund_size_aud_millions': _safe_float(cell('fum')),
            'fund_flow_1m':  _safe_float(cell('flow_1m')),
            'fund_flow_3m':  _safe_float(cell('flow_3m')),
            'fund_flow_1y':  _safe_float(cell('flow_1y')),
            'management_fee': _safe_float(cell('mer')),
            # Returns are stored as decimal fractions in the ASX ETP report
            'return_1m': _from_decimal_pct(_safe_float(cell('return_1m'))),
            'return_1y': _from_decimal_pct(_safe_float(cell('return_1y'))),
            'return_3y': _from_decimal_pct(_safe_float(cell('return_3y'))),
            'return_5y': _from_decimal_pct(_safe_float(cell('return_5y'))),
            # Price from "Last ($)" column
            'current_price': _safe_float(cell('price')),
            # Yield also stored as decimal fraction
            'distribution_yield': _from_decimal_pct(_safe_float(cell('yield'))),
            'data_source': 'asx_report',
            'asx_url': f'https://www2.asx.com.au/markets/etp/{code}',
        }

        # Remove None values
        etf = {k: v for k, v in etf.items() if v is not None}
        if 'code' in etf:
            results.append(etf)

    return results


def scrape_asx_report(db_path=None) -> int:
    """
    Main entry point: download ASX monthly report, parse all ETF sheets,
    upsert into database.  Returns count of ETFs processed.
    """
    started = datetime.utcnow()
    source = 'asx_report'

    try:
        import openpyxl
    except ImportError:
        logger.error("openpyxl is required. Install with: pip install openpyxl")
        return 0

    xlsx_path = _download_report()
    if not xlsx_path:
        logger.error("Failed to download ASX investment products report")
        conn = get_connection(db_path)
        log_scrape(conn, source, 'error', error='Download failed',
                   duration_secs=(datetime.utcnow() - started).total_seconds(),
                   started_at=started.isoformat())
        conn.close()
        return 0

    logger.info(f"Parsing {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    # Sheets to always skip — mFunds, LICs, A-REITs, and Infrastructure are
    # not ETFs.  Matching 'fund' alone was too broad (catches 'mFund List').
    SKIP_KEYWORDS = ['mfund', 'lic ', 'lic\n', 'a-reit', 'reit', 'infra']

    all_etfs = []
    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        name_lower = ws_name.lower()
        # Skip non-ETF sheets explicitly
        if any(kw in name_lower for kw in SKIP_KEYWORDS):
            logger.debug(f"  Skipping non-ETF sheet '{ws_name}'")
            continue
        # Only process sheets that look like ETF/ETP listings
        if any(kw in name_lower for kw in ['etf', 'etp', 'exchange traded']):
            etfs = _parse_etf_sheet(ws)
            if etfs:
                logger.info(f"  Sheet '{ws_name}': found {len(etfs)} ETFs")
                all_etfs.extend(etfs)

    # If no ETF-specific sheets found, try all non-excluded sheets
    if not all_etfs:
        logger.info("No ETF-specific sheets found, trying all non-mFund/LIC sheets...")
        for ws_name in wb.sheetnames:
            ws = wb[ws_name]
            name_lower = ws_name.lower()
            if any(kw in name_lower for kw in SKIP_KEYWORDS):
                continue
            etfs = _parse_etf_sheet(ws)
            if etfs:
                logger.info(f"  Sheet '{ws_name}': found {len(etfs)} ETFs")
                all_etfs.extend(etfs)

    wb.close()

    # Deduplicate by code (keep first occurrence which usually has more data)
    seen = set()
    unique_etfs = []
    for etf in all_etfs:
        if etf['code'] not in seen:
            seen.add(etf['code'])
            unique_etfs.append(etf)

    logger.info(f"Total unique ETFs from ASX report: {len(unique_etfs)}")

    conn = get_connection(db_path)
    upsert_etfs(conn, unique_etfs)
    update_issuer_stats(conn)

    duration = (datetime.utcnow() - started).total_seconds()
    log_scrape(conn, source, 'success', records_affected=len(unique_etfs),
               duration_secs=duration, started_at=started.isoformat())
    conn.close()

    return len(unique_etfs)
