"""
Per-issuer website scrapers.

Each function scrapes an issuer's website for ETF data that isn't available
from ASX/Cboe sources: holdings, sector allocations, returns, fees, etc.
"""

import re
import io
import json
import html as html_module
import logging
from datetime import datetime, date

from scrapers.config import ISSUER_URLS, SPDR_AU_FUNDS, VANGUARD_AU_FUNDS, VANGUARD_AU_PORT_IDS, normalise_asset_class, CXA_ISSUER_URLS
from scrapers.base_scraper import fetch, fetch_json
from scrapers.db_writer import (
    get_connection, upsert_etf, upsert_holdings, upsert_sectors,
    log_scrape, update_issuer_stats,
)

logger = logging.getLogger(__name__)


def _safe_float(val) -> float | None:
    if val is None:
        return None
    try:
        s = str(val).replace(',', '').replace('$', '').replace('%', '').strip()
        if s in ('', '-', 'N/A', 'n/a'):
            return None
        return float(s)
    except (ValueError, TypeError):
        return None


def _find_json_in_script(soup, key_hints: list[str]) -> dict | list | None:
    """Search <script> tags for inline JSON containing any of the given key strings."""
    for tag in soup.find_all('script'):
        text = tag.string or ''
        if not any(h in text for h in key_hints):
            continue
        # Try to extract the largest JSON object/array in the block
        for match in re.finditer(r'(\{[\s\S]{20,}\}|\[[\s\S]{20,}\])', text):
            try:
                return json.loads(match.group(0))
            except (ValueError, json.JSONDecodeError):
                continue
    return None


# ====================================================================
# BetaShares
# ====================================================================

def _scrape_betashares_fund_page(conn, slug: str, soup) -> str | None:
    """
    Parse a BetaShares individual fund page.
    Returns the ASX code, or None if not found.
    Title format: "NDQ ASX | Nasdaq 100 ETF | Betashares"
    Holdings table rows: <th>COMPANY NAME</th><td>weight</td>
    """
    # ASX code from <title> — handles multiple formats:
    #   "NDQ ASX | Nasdaq 100 ETF | Betashares"  (old)
    #   "ASX A200 | Australia 200 ETF | Betashares"  (new)
    #   "AGVT ETF | Australian Government Bond ETF | Betashares"  (new alt)
    title = soup.title.string if soup.title else ''
    code_match = (
        re.match(r'^([A-Z0-9]{2,6})\s+ASX\s*\|', title) or   # old: CODE ASX |
        re.match(r'^ASX\s+([A-Z0-9]{2,6})\s*[\|\s]', title) or  # new: ASX CODE |
        re.match(r'^([A-Z0-9]{2,6})\s+ETF\s*\|', title)          # new alt: CODE ETF |
    )
    if not code_match:
        return None
    code = code_match.group(1)

    # Fund name from title
    name_match = re.search(r'(?:ASX\s+)?[A-Z0-9]+\s+(?:ASX\s+)?\|\s+(.+?)\s+\|\s+', title)
    name = name_match.group(1) if name_match else None

    # MER from the "Management fee and cost" table row
    mer = None
    for table in soup.find_all('table'):
        headers = [th.get_text(strip=True).lower() for th in table.find_all('th')]
        if 'management fee and cost** (p.a.)' in headers or 'management fee and costs (p.a.)' in headers:
            # MER is in the first data td in this table
            for row in table.find_all('tr')[1:]:
                tds = [td.get_text(strip=True) for td in row.find_all('td')]
                if tds:
                    val = _safe_float(tds[0])
                    if val and 0 < val < 5:
                        mer = val
                    break
            break

    # Benchmark from the "Index" table row
    benchmark = None
    for table in soup.find_all('table'):
        for row in table.find_all('tr'):
            cells = row.find_all(['td', 'th'])
            if len(cells) >= 2:
                label = cells[0].get_text(strip=True).lower()
                if label in ('index', 'underlying index', 'benchmark index', 'benchmark'):
                    val = cells[1].get_text(strip=True)
                    if val and len(val) > 3:
                        benchmark = val
                        break
        if benchmark:
            break

    # Holdings — rows are <tr><th>COMPANY NAME</th><td>weight%</td></tr>
    holdings = []
    for table in soup.find_all('table'):
        first_row_headers = [th.get_text(strip=True).lower() for th in table.find_all('tr')[0].find_all('th')] if table.find_all('tr') else []
        if 'name' in first_row_headers and 'weight (%)' in first_row_headers:
            for row in table.find_all('tr')[1:]:
                th_els = row.find_all('th')
                td_els = row.find_all('td')
                if th_els and td_els:
                    h_name = th_els[0].get_text(strip=True).title()
                    weight = _safe_float(td_els[0].get_text(strip=True))
                    if h_name and weight:
                        holdings.append({'name': h_name, 'weight_pct': weight})
            break

    etf = {
        'code': code,
        'name': name,
        'issuer': 'BetaShares',
        'expense_ratio': mer,
        'management_fee': mer,
        'benchmark': benchmark,
        'data_source': 'betashares',
        'issuer_url': f"https://www.betashares.com.au/fund/{slug}/",
    }
    etf = {k: v for k, v in etf.items() if v is not None}
    upsert_etf(conn, etf)

    if holdings:
        # Deduplicate by name — merge weights when the same name appears twice
        # (e.g. Alphabet Inc appears as both GOOGL and GOOG share classes)
        merged: dict[str, dict] = {}
        for h in holdings:
            name = h['name']
            if name in merged:
                merged[name]['weight_pct'] = (merged[name]['weight_pct'] or 0) + (h['weight_pct'] or 0)
            else:
                merged[name] = h
        upsert_holdings(conn, code, list(merged.values()), commit=False)

    return code


def scrape_betashares(db_path=None) -> int:
    """Scrape BetaShares fund list (/fund/) and individual fund pages for MER + holdings."""
    started = datetime.utcnow()
    source = 'betashares'
    updated = 0

    try:
        from bs4 import BeautifulSoup
    except ImportError:
        logger.error("beautifulsoup4 required for issuer scrapers")
        return 0

    conn = get_connection(db_path)
    try:
        fund_list_url = ISSUER_URLS['BetaShares']['fund_list']

        # Step 1: Get all fund slugs from the listing page
        resp = fetch(fund_list_url)
        if not resp:
            log_scrape(conn, source, 'error', error='Failed to fetch fund list',
                       duration_secs=(datetime.utcnow() - started).total_seconds(),
                       started_at=started.isoformat())
            return 0

        soup = BeautifulSoup(resp.text, 'html.parser')
        slugs: list[str] = []
        seen_slugs: set[str] = set()
        for a in soup.find_all('a', href=True):
            href = a['href']
            m = re.match(r'https://www\.betashares\.com\.au/fund/([^/?#]+)/?$', href)
            if m:
                slug = m.group(1)
                if slug not in seen_slugs:
                    seen_slugs.add(slug)
                    slugs.append(slug)

        logger.info(f"BetaShares: found {len(slugs)} fund slugs")

        # Step 2: Scrape each individual fund page
        for slug in slugs:
            fund_url = f"https://www.betashares.com.au/fund/{slug}/"
            resp = fetch(fund_url)
            if not resp:
                continue
            page_soup = BeautifulSoup(resp.text, 'html.parser')
            try:
                code = _scrape_betashares_fund_page(conn, slug, page_soup)
            except Exception as e:
                logger.warning(f"BetaShares: error scraping {slug}: {e}")
                conn.rollback()
                continue
            if code:
                updated += 1

        conn.commit()
        duration = (datetime.utcnow() - started).total_seconds()
        log_scrape(conn, source, 'success' if updated else 'no_data',
                   records_affected=updated, duration_secs=duration,
                   started_at=started.isoformat())
        logger.info(f"BetaShares: updated {updated} ETFs")
        return updated
    finally:
        conn.close()


# ====================================================================
# VanEck
# ====================================================================

def _scrape_vaneck_snapshot(conn, code: str, snapshot_url: str) -> None:
    """
    Scrape a VanEck fund snapshot page for MER, inception date, and holdings.
    Primary path: extract the embedded FundDatasetBlock/Get/ JSON API URL from the
    page HTML and call it directly (returns structured JSON with holdings + metadata).
    Fallback: BeautifulSoup table parsing for any missing fields.
    """
    try:
        from bs4 import BeautifulSoup
    except ImportError:
        return

    resp = fetch(snapshot_url)
    if not resp:
        return

    update = {'code': code}
    holdings = []

    # --- Primary: FundDatasetBlock JSON API ---
    # VanEck embeds a URL like:
    #   FundDatasetBlock/Get/?blockId=352386&pageId=248665&ticker=QUALAU
    # in the page HTML which returns full fund metadata + holdings as JSON.
    api_match = re.search(
        r'FundDatasetBlock/Get/\?blockId=(\d+)&pageId=(\d+)&ticker=(\w+)',
        resp.text
    )
    if api_match:
        block_id, page_id, ticker = api_match.groups()
        api_url = (
            f'https://www.vaneck.com.au/Main/FundDatasetBlock/Get/'
            f'?blockId={block_id}&pageId={page_id}&ticker={ticker}'
        )
        api_resp = fetch(api_url)
        if api_resp:
            try:
                data = api_resp.json()
                # Management fee — stored as "0.40%" string
                fee_str = data.get('Management fee (p.a.)', '') or ''
                fee_val = _safe_float(re.sub(r'[^0-9.]', '', fee_str))
                if fee_val and 0 < fee_val < 5:
                    update['expense_ratio'] = fee_val
                # Inception date
                inc = (data.get('Inception Date') or '').strip()
                if inc:
                    update['inception_date'] = inc
                # Holdings — HoldingsList is a list of date-bucketed snapshots;
                # use the first (most recent) entry.
                for entry in data.get('HoldingsList', []):
                    for h in entry.get('Holdings', []):
                        name = (h.get('HoldingName') or '').strip()
                        label = (h.get('Label') or '').strip()  # e.g. "META US"
                        weight = _safe_float(h.get('Weight'))
                        if name and weight is not None and 0 < weight < 100:
                            holdings.append({
                                'name': name,
                                'ticker': label,
                                'weight_pct': weight,
                            })
                    if holdings:
                        break  # only need the most recent snapshot
            except (ValueError, KeyError, AttributeError):
                pass

    # --- Fallback: BeautifulSoup table parsing ---
    soup = BeautifulSoup(resp.text, 'html.parser')

    if 'expense_ratio' not in update:
        for label in soup.find_all(string=re.compile(r'Management\s+(Fee|Cost|Expense|MER)', re.I)):
            parent = label.parent
            for candidate in [parent.find_next_sibling(), parent.find_next('td')]:
                if candidate:
                    text = re.sub(r'[^0-9.]', '', candidate.get_text())
                    val = _safe_float(text)
                    if val and 0 < val < 5:
                        update['expense_ratio'] = val
                        break
            if 'expense_ratio' in update:
                break

    if 'inception_date' not in update:
        for label in soup.find_all(string=re.compile(r'Inception\s+Date', re.I)):
            parent = label.parent
            for candidate in [parent.find_next_sibling(), parent.find_next('td')]:
                if candidate:
                    text = candidate.get_text(strip=True)
                    if re.match(r'\d{1,2}[/\-]\w+[/\-]\d{2,4}', text) or re.match(r'\d{4}-\d{2}-\d{2}', text):
                        update['inception_date'] = text
                        break
            if 'inception_date' in update:
                break

    # Benchmark — two patterns:
    # 1. Structured: <h4>Underlying Index:</h4><p>{name}</p>
    # 2. Prose: "the benchmark, the {Index Name}"
    if 'benchmark' not in update:
        for el in soup.find_all(string=re.compile(r'Underlying\s+Index', re.I)):
            parent = el.parent
            for sib in parent.next_siblings:
                sib_text = sib.get_text(strip=True) if hasattr(sib, 'get_text') else ''
                if sib_text and len(sib_text) > 3:
                    update['benchmark'] = sib_text
                    break
            if 'benchmark' in update:
                break
    if 'benchmark' not in update:
        bench_m = re.search(
            r'the benchmark,\s+the\s+([A-Z][^\n<.]{5,100}Index)',
            resp.text,
        )
        if bench_m:
            bench_val = bench_m.group(1).strip()
            if len(bench_val) > 5 and 'Index' in bench_val:
                update['benchmark'] = bench_val

    # Returns table fallback
    for table in soup.find_all('table'):
        headers = [th.get_text(strip=True).lower() for th in table.find_all('th')]
        has_returns = any(
            any(h in hdr for h in ['1 yr', '1yr', '1 year', '3 yr', '3yr', 'ytd'])
            for hdr in headers
        )
        if not has_returns:
            continue
        for row in table.find_all('tr')[1:]:
            cells = [td.get_text(strip=True) for td in row.find_all('td')]
            if not cells:
                continue
            if any(kw in cells[0].lower() for kw in ('fund', 'nav', 'etf')):
                col_map = {}
                for i, h in enumerate(headers):
                    if 'month' in h or '1m' in h:
                        col_map[i] = 'return_1m'
                    elif '1' in h and ('yr' in h or 'year' in h):
                        col_map[i] = 'return_1y'
                    elif '3' in h and ('yr' in h or 'year' in h):
                        col_map[i] = 'return_3y'
                    elif '5' in h and ('yr' in h or 'year' in h):
                        col_map[i] = 'return_5y'
                for i, field in col_map.items():
                    if i < len(cells):
                        update[field] = _safe_float(cells[i])
                break

    if not holdings:
        for table in soup.find_all('table'):
            headers = [th.get_text(strip=True).lower() for th in table.find_all('th')]
            if not (any(h in ' '.join(headers) for h in ['holding', 'security', 'name']) and
                    any(h in ' '.join(headers) for h in ['weight', '%', 'allocation'])):
                continue
            for row in table.find_all('tr')[1:]:
                cells = [td.get_text(strip=True) for td in row.find_all('td')]
                if len(cells) < 2:
                    continue
                name = cells[0] if cells[0] and cells[0] != '-' else None
                weight = None
                for cell in reversed(cells):
                    w = _safe_float(cell)
                    if w and 0 < w < 100:
                        weight = w
                        break
                if name and weight:
                    holdings.append({'name': name, 'weight_pct': weight})
            if holdings:
                break

    if len(update) > 1:
        upsert_etf(conn, update)
    if holdings:
        # Deduplicate by name (merge weights for same-name duplicates e.g. Alphabet Inc)
        merged: dict[str, dict] = {}
        for h in holdings:
            name = h['name']
            if name in merged:
                merged[name]['weight_pct'] = (merged[name]['weight_pct'] or 0) + (h.get('weight_pct') or 0)
            else:
                merged[name] = dict(h)
        upsert_holdings(conn, code, list(merged.values()), commit=False)


def scrape_vaneck(db_path=None) -> int:
    """Scrape VanEck Australia ETF list and individual snapshot pages."""
    started = datetime.utcnow()
    source = 'vaneck'
    updated = 0

    try:
        from bs4 import BeautifulSoup
    except ImportError:
        return 0

    conn = get_connection(db_path)

    # Fund list — VanEck /etf/ lists all ETFs via snapshot links
    resp = fetch('https://www.vaneck.com.au/etf/')
    if not resp:
        log_scrape(conn, source, 'error', error='Failed to fetch fund list',
                   duration_secs=(datetime.utcnow() - started).total_seconds(),
                   started_at=started.isoformat())
        conn.close()
        return 0

    soup = BeautifulSoup(resp.text, 'html.parser')

    # Extract ETFs from links like /etf/equity/qual/snapshot
    seen: dict[str, str] = {}  # code -> snapshot URL
    for link in soup.find_all('a', href=True):
        href = link['href']
        match = re.match(r'^/etf/[^/]+/([a-z0-9]{2,6})/snapshot/?$', href, re.I)
        if not match:
            continue

        code = match.group(1).upper()
        if code in seen:
            continue

        text = link.get_text(strip=True)
        name = None
        if text.startswith(code):
            name = text[len(code):].strip()
        elif text:
            name = text

        category_segment = href.split('/')[2] if len(href.split('/')) > 2 else None
        asset_class = None
        if category_segment:
            cat_map = {
                'equity': 'International Equities',
                'fixed-income': 'Fixed Income',
                'income': 'Fixed Income',
                'alternatives': 'Alternatives',
                'multi-asset': 'Diversified',
                'australian-equities': 'Australian Equities',
                'australian': 'Australian Equities',
            }
            asset_class = cat_map.get(category_segment)

        snapshot_url = f"https://www.vaneck.com.au{href}"
        seen[code] = snapshot_url

        etf = {
            'code': code,
            'name': name,
            'issuer': 'VanEck',
            'asset_class': asset_class,
            'data_source': 'vaneck',
            'issuer_url': snapshot_url,
        }
        etf = {k: v for k, v in etf.items() if v is not None}
        upsert_etf(conn, etf)
        updated += 1

    conn.commit()

    # Scrape individual snapshot pages for MER, returns, holdings (top 30 by priority)
    if seen:
        # Prioritise ETFs already in DB with FUM ranking
        ranked = conn.execute(
            "SELECT code FROM etfs WHERE issuer = 'VanEck' AND code IN ({}) ORDER BY rank_by_fum LIMIT 30".format(
                ','.join('?' * len(seen))
            ),
            list(seen.keys())
        ).fetchall()
        priority_codes = [r[0] for r in ranked]
        # Add any remaining codes not yet ranked
        remaining = [c for c in seen if c not in priority_codes]
        codes_to_scrape = (priority_codes + remaining)[:30]

        detail_updated = 0
        for code in codes_to_scrape:
            try:
                _scrape_vaneck_snapshot(conn, code, seen[code])
                detail_updated += 1
            except Exception as e:
                logger.warning(f"VanEck: error scraping {code}: {e}")
                conn.rollback()

        conn.commit()
        logger.info(f"VanEck: scraped detail pages for {detail_updated} ETFs")

    duration = (datetime.utcnow() - started).total_seconds()
    try:
        log_scrape(conn, source, 'success' if updated else 'no_data',
                   records_affected=updated, duration_secs=duration,
                   started_at=started.isoformat())
    finally:
        conn.close()
    logger.info(f"VanEck: updated {updated} ETFs")
    return updated


# ====================================================================
# Vanguard
# ====================================================================

def scrape_vanguard(db_path=None) -> int:
    """
    Seed Vanguard Australia ETF data from the known fund list.
    (Vanguard AU website is an Angular SPA — no accessible server-side API.)
    Upserts the known fund list and preserves any richer data already in DB.
    """
    started = datetime.utcnow()
    source = 'vanguard'
    updated = 0

    conn = get_connection(db_path)

    # Seed from known list (preserves existing data via COALESCE upsert)
    for code, name in VANGUARD_AU_FUNDS.items():
        port_id = VANGUARD_AU_PORT_IDS.get(code)
        issuer_url = (
            f"https://www.vanguard.com.au/personal/invest-with-us/etf?portId={port_id}"
            if port_id else None
        )
        etf = {
            'code': code,
            'name': name,
            'issuer': 'Vanguard',
            'exchange': 'ASX',
            'data_source': 'vanguard',
        }
        if issuer_url:
            etf['issuer_url'] = issuer_url
        upsert_etf(conn, etf)
        updated += 1

    # Fix issuer_url for all Vanguard ETFs already in the DB using the portId mapping
    for code, port_id in VANGUARD_AU_PORT_IDS.items():
        conn.execute("""
            UPDATE etfs
            SET issuer_url = ?
            WHERE code = ? AND issuer = 'Vanguard'
        """, (f"https://www.vanguard.com.au/personal/invest-with-us/etf?portId={port_id}", code))

    conn.commit()
    duration = (datetime.utcnow() - started).total_seconds()
    log_scrape(conn, source, 'success', records_affected=updated,
               duration_secs=duration, started_at=started.isoformat())
    conn.close()
    logger.info(f"Vanguard: seeded {updated} ETFs from known fund list")
    return updated


# ====================================================================
# iShares (BlackRock)
# ====================================================================

def _scrape_ishares_product_page(conn, code: str, product_url: str) -> bool:
    """
    Fetch all holdings for a single iShares ETF from its BlackRock product page.

    BlackRock product pages embed AJAX endpoint URLs of the form:
        /au/products/{productId}/fund/{ajaxId}.ajax
    We find the holdings-specific endpoint (identified by a 'holdings' filename
    in the download link) and call it with tab=all&fileType=json.

    JSON response shape:
        {"aaData": [[ticker, name, sector, asset_class, market_value, weight,
                     country, currency], ...]}
    where weight is either a plain float or {"raw": 10.83, "display": "10.83"}.

    Returns True if any holdings were stored.
    """
    resp = fetch(product_url)
    if not resp:
        return False

    # Find the holdings AJAX base path from the CSV download link embedded in the page.
    # Example: /au/products/251852/fund/1478358644060.ajax?fileType=csv&fileName=IOZ_holdings
    holdings_match = re.search(
        r'(/au/products/\d+/fund/(\d+)\.ajax)[^"\']*fileName=[^"\']*[Hh]olding',
        resp.text,
    )
    if not holdings_match:
        logger.debug(f"iShares {code}: holdings ajax URL not found in product page")
        return False

    base_ajax_path = holdings_match.group(1)
    json_url = f'https://www.blackrock.com{base_ajax_path}?tab=all&fileType=json'

    # BlackRock returns UTF-8 with BOM — strip it before JSON parsing
    json_resp = fetch(json_url)
    if not json_resp:
        logger.debug(f"iShares {code}: no response from holdings URL")
        return False
    try:
        data = json.loads(json_resp.text.lstrip('\ufeff'))
    except (ValueError, json.JSONDecodeError):
        logger.debug(f"iShares {code}: invalid JSON in holdings response")
        return False
    if 'aaData' not in data:
        logger.debug(f"iShares {code}: no aaData in holdings response")
        return False

    # Row format (15 columns):
    #   0:ticker  1:name  2:sector  3:asset_class  4:market_value  5:weight
    #   6:notional  7:shares  8:security_code  9:ISIN  10:exchange_code
    #   11:price  12:country  13:exchange  14:currency
    holdings = []
    for row in data['aaData']:
        if len(row) < 6:
            continue
        ticker = str(row[0]).strip() or None
        name   = str(row[1]).strip() or None
        sector = str(row[2]).strip() or None
        # Weight is at index 5: {"raw": 10.82906, "display": "10.83"} or plain value
        weight_raw = row[5]
        if isinstance(weight_raw, dict):
            weight = _safe_float(weight_raw.get('raw') or weight_raw.get('display'))
        else:
            weight = _safe_float(weight_raw)
        # Country position varies by fund type (equity: index 12, fixed income: index 13+)
        # Scan from index 12 for the first plain string that looks like a country name
        country = None
        for idx in range(12, min(16, len(row))):
            val = row[idx]
            if isinstance(val, str) and val not in ('-', 'N/A', '') and not val.isdigit():
                country = val.strip()
                break

        if not name or name == '-':
            continue
        if weight is None or not (0 < weight < 100):
            continue

        holdings.append({
            'ticker':     ticker if ticker and ticker != '-' else None,
            'name':       name,
            'sector':     sector if sector and sector not in ('-', 'N/A') else None,
            'weight_pct': weight,
            'country':    country if country and country not in ('-', 'N/A') else None,
        })

    if not holdings:
        return False

    # Deduplicate by name (merge weights for same-name entries e.g. dual-listed shares)
    merged: dict[str, dict] = {}
    for h in holdings:
        key = h['name']
        if key in merged:
            merged[key]['weight_pct'] = (merged[key]['weight_pct'] or 0) + (h['weight_pct'] or 0)
        else:
            merged[key] = dict(h)

    upsert_holdings(conn, code, list(merged.values()), commit=False)

    # Aggregate sector allocations from holdings (skip cash/derivatives for sector summary)
    sector_weights: dict[str, float] = {}
    for h in merged.values():
        s = h.get('sector')
        w = h.get('weight_pct') or 0
        if s and s != 'Cash and/or Derivatives':
            sector_weights[s] = sector_weights.get(s, 0) + w
    if sector_weights:
        sectors = [{'sector': s, 'weight_pct': round(w, 6)} for s, w in sector_weights.items()]
        upsert_sectors(conn, code, sectors, commit=False)

    logger.debug(f"iShares {code}: stored {len(merged)} holdings, {len(sector_weights)} sectors")
    return True


def scrape_ishares(db_path=None) -> int:
    """Scrape iShares/BlackRock Australia ETF data."""
    started = datetime.utcnow()
    source = 'ishares'
    updated = 0

    conn = get_connection(db_path)

    # Try multiple known API patterns for BlackRock AU
    api_attempts = [
        ('https://www.blackrock.com/au/individual/products/product-list?productView=etf&iShares=true&country=au&domicileCountry=au&type=ishares&dataType=fund', {}),
        ('https://www.blackrock.com/au/individual/products/fund-list', {'Accept': 'application/json'}),
        ('https://www.blackrock.com/cache/api/fund/getfundlist/?productView=etf&country=au&iShares=true', {}),
    ]

    for api_url, extra_headers in api_attempts:
        headers = {'Accept': 'application/json', 'User-Agent': 'Mozilla/5.0'}
        headers.update(extra_headers)
        data = fetch_json(api_url, headers=headers)
        if not data:
            continue

        funds_raw = None
        if isinstance(data, list):
            funds_raw = data
        elif isinstance(data, dict):
            for key in ('data', 'funds', 'products', 'fundData', 'aaData'):
                if key in data:
                    funds_raw = data[key]
                    break

        if not funds_raw:
            continue

        for fund in funds_raw:
            if not isinstance(fund, dict):
                continue
            code = (fund.get('localExchangeTicker') or fund.get('ticker') or
                    fund.get('fundCode') or fund.get('asx_code') or '').strip().upper()
            if not code or not re.match(r'^[A-Z0-9]{2,6}$', code) or code.isdigit():
                continue

            etf = {
                'code': code,
                'name': fund.get('fundName') or fund.get('name'),
                'issuer': 'iShares',
                'expense_ratio': _safe_float(fund.get('totalExpenseRatio') or fund.get('mer') or fund.get('managementFee')),
                'asset_class': normalise_asset_class(fund.get('assetClass') or fund.get('asset_class')),
                'benchmark': fund.get('benchmark') or fund.get('benchmarkName'),
                'inception_date': fund.get('inceptionDate') or fund.get('fundInceptionDate'),
                'fund_size_aud_millions': _safe_float(fund.get('totalNetAssets') or fund.get('aum')),
                'return_1y': _safe_float(fund.get('return1yr') or fund.get('return1y') or fund.get('annualReturn1yr')),
                'return_3y': _safe_float(fund.get('return3yr') or fund.get('return3y')),
                'return_5y': _safe_float(fund.get('return5yr') or fund.get('return5y')),
                'data_source': 'ishares',
            }
            etf = {k: v for k, v in etf.items() if v is not None}
            upsert_etf(conn, etf)
            updated += 1

        if updated:
            break  # Stop trying other API endpoints

    # HTML fallback — parse fund list page and extract inline JSON
    if updated == 0:
        try:
            from bs4 import BeautifulSoup
        except ImportError:
            pass
        else:
            fund_list_url = 'https://www.blackrock.com/au/individual/products/investment-funds#categoryId=702702&tab=overview'
            resp = fetch(fund_list_url)
            if resp:
                soup = BeautifulSoup(resp.text, 'html.parser')

                # BlackRock sometimes embeds fund data in a JS variable
                json_data = _find_json_in_script(soup, ['localExchangeTicker', 'fundName', 'iShares'])
                if json_data:
                    funds = json_data if isinstance(json_data, list) else json_data.get('data', [])
                    for fund in funds:
                        if not isinstance(fund, dict):
                            continue
                        code = (fund.get('localExchangeTicker') or fund.get('ticker') or '').strip().upper()
                        if not re.match(r'^[A-Z0-9]{2,6}$', code) or code.isdigit():
                            continue
                        etf = {
                            'code': code,
                            'name': fund.get('fundName') or fund.get('name'),
                            'issuer': 'iShares',
                            'expense_ratio': _safe_float(fund.get('totalExpenseRatio') or fund.get('mer')),
                            'asset_class': normalise_asset_class(fund.get('assetClass')),
                            'data_source': 'ishares',
                        }
                        etf = {k: v for k, v in etf.items() if v is not None}
                        upsert_etf(conn, etf)
                        updated += 1
                else:
                    # Last resort: extract ticker-like codes from links to fund pages.
                    # Require at least one letter — BlackRock numeric product IDs like
                    # /products/251852/ must not be treated as ASX ticker codes.
                    for link in soup.find_all('a', href=True):
                        href = link['href']
                        if '/products/' not in href and '/funds/' not in href:
                            continue
                        # iShares AU product page patterns — [A-Z] anchor prevents
                        # matching pure-numeric BlackRock internal product IDs.
                        for pattern in [
                            r'/products/([A-Z][A-Z0-9]{1,5})/',
                            r'ticker=([A-Z][A-Z0-9]{1,5})',
                            r'/etf/([A-Z][A-Z0-9]{1,5})$',
                        ]:
                            match = re.search(pattern, href, re.IGNORECASE)
                            if match:
                                code = match.group(1).upper()
                                name = link.get_text(strip=True)
                                if name and 3 < len(name) < 100:
                                    etf = {
                                        'code': code,
                                        'name': name,
                                        'issuer': 'iShares',
                                        'data_source': 'ishares',
                                    }
                                    upsert_etf(conn, etf)
                                    updated += 1
                                break

    conn.commit()

    # ---- Phase 2: holdings from individual product pages ----
    holdings_count = 0
    try:
        cur = conn.cursor()
        cur.execute(
            "SELECT code, issuer_url FROM etfs "
            "WHERE issuer = 'iShares' "
            "AND issuer_url LIKE 'https://www.blackrock.com/au/products/%'"
        )
        product_rows = cur.fetchall()
    except Exception as e:
        logger.warning(f"iShares: could not query product URLs: {e}")
        product_rows = []

    for row in product_rows:
        code_h, url_h = row['code'], row['issuer_url']
        try:
            if _scrape_ishares_product_page(conn, code_h, url_h):
                holdings_count += 1
        except Exception as e:
            logger.warning(f"iShares holdings {code_h}: {e}")
        conn.commit()

    logger.info(f"iShares: fetched holdings for {holdings_count}/{len(product_rows)} ETFs")

    duration = (datetime.utcnow() - started).total_seconds()
    log_scrape(conn, source, 'success' if (updated or holdings_count) else 'no_data',
               records_affected=updated + holdings_count, duration_secs=duration,
               started_at=started.isoformat())
    conn.close()
    logger.info(f"iShares: updated {updated} ETFs")
    return updated


# ====================================================================
# SSGA Holdings (shared by SPDR + StateStreet)
# ====================================================================

def _parse_ssga_holdings_excel(content: bytes) -> list[dict]:
    """
    Parse an SSGA holdings-daily Excel file.
    URL pattern: https://www.ssga.com/library-content/products/fund-data/etfs/apac/holdings-daily-au-en-{ticker}.xlsx
    Structure:
      Row 0-2: metadata (Fund Name, Ticker, Holdings date)
      Row 4:   header (ISIN, SEDOL, Ticker, Name, Currency, Shares, Weight(%), Country, Price, Sector, Industry)
      Row 5+:  holdings rows
    Returns list of {name, ticker, isin, weight_pct, sector, country}.
    """
    try:
        import openpyxl
    except ImportError:
        logger.error("openpyxl required for SSGA holdings parsing")
        return []

    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        logger.warning(f"SSGA Excel parse error: {e}")
        return []

    ws = wb.active
    header_row = None
    col = {}
    holdings = []

    for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
        # Find header row (contains 'Weight' or 'Name')
        if header_row is None:
            row_lower = [str(v or '').strip().lower() for v in row]
            if any('weight' in c for c in row_lower):
                header_row = row_idx
                for i, h in enumerate(row_lower):
                    if 'name' in h and 'fund' not in h:
                        col.setdefault('name', i)
                    elif 'ticker' in h or 'symbol' in h:
                        col.setdefault('ticker', i)
                    elif 'isin' in h:
                        col.setdefault('isin', i)
                    elif 'weight' in h:
                        col.setdefault('weight', i)
                    elif 'sector' in h and 'industry' not in h:
                        col.setdefault('sector', i)
                    elif 'country' in h or 'trade country' in h:
                        col.setdefault('country', i)
            continue

        if header_row is None or not col:
            continue

        def gcol(key):
            idx = col.get(key)
            return row[idx] if idx is not None and idx < len(row) else None

        name = str(gcol('name') or '').strip()
        if not name or name.lower() in ('cash', '-', 'n/a', ''):
            continue

        weight = _safe_float(gcol('weight'))
        if weight is None or weight <= 0:
            continue

        holdings.append({
            'name':       name,
            'ticker':     str(gcol('ticker') or '').strip() or None,
            'isin':       str(gcol('isin') or '').strip() or None,
            'weight_pct': weight,
            'sector':     str(gcol('sector') or '').strip() or None,
            'country':    str(gcol('country') or '').strip() or None,
        })

    return holdings


def _scrape_ssga_holdings(conn, code: str) -> bool:
    """
    Download and upsert SSGA holdings for a single ETF code.
    Returns True if holdings were found and saved.
    """
    url = f'https://www.ssga.com/library-content/products/fund-data/etfs/apac/holdings-daily-au-en-{code.lower()}.xlsx'
    resp = fetch(url)
    if not resp or resp.status_code != 200:
        return False

    rows = _parse_ssga_holdings_excel(resp.content)
    if not rows:
        return False

    # Deduplicate by name
    merged: dict[str, dict] = {}
    for h in rows:
        n = h['name']
        if n in merged:
            merged[n]['weight_pct'] = (merged[n]['weight_pct'] or 0) + (h['weight_pct'] or 0)
        else:
            merged[n] = h

    upsert_holdings(conn, code, list(merged.values()), commit=False)

    # Aggregate sector allocations
    sector_weights: dict[str, float] = {}
    for h in merged.values():
        s = h.get('sector')
        if s and s.lower() not in ('cash', 'cash and/or derivatives', '-', 'n/a', ''):
            sector_weights[s] = sector_weights.get(s, 0) + (h.get('weight_pct') or 0)
    if sector_weights:
        sectors = [{'sector': s, 'weight_pct': round(w, 4)} for s, w in sector_weights.items()]
        upsert_sectors(conn, code, sectors, commit=False)

    return True


# ====================================================================
# SPDR (State Street)
# ====================================================================

def scrape_spdr(db_path=None) -> int:
    """
    Scrape SPDR/State Street Australia ETF data from individual fund pages.
    (SSGA fund finder is JS-rendered; individual pages serve static HTML with JSON-LD.)
    Uses SPDR_AU_FUNDS from config as the known fund index.
    """
    started = datetime.utcnow()
    source = 'spdr'
    updated = 0

    try:
        from bs4 import BeautifulSoup
    except ImportError:
        return 0

    conn = get_connection(db_path)
    base_url = 'https://www.ssga.com/au/en_gb/intermediary/etfs/funds/'

    for code, slug in SPDR_AU_FUNDS.items():
        url = base_url + slug
        resp = fetch(url)
        if not resp:
            continue

        soup = BeautifulSoup(resp.text, 'html.parser')

        # Name and ticker from JSON-LD schema
        name = None
        for script in soup.find_all('script', type='application/ld+json'):
            try:
                ld = json.loads(script.string.strip())
                raw_name = ld.get('name', '')
                # Strip HTML entities and trademark symbols
                name = re.sub(r'[®™]|&\w+;', '', raw_name).strip() or None
                # Verify this page is for the right ticker
                ticker_symbol = ld.get('tickerSymbol', '')
                if code not in ticker_symbol:
                    name = None  # Wrong page
                break
            except (ValueError, json.JSONDecodeError):
                continue

        if not name:
            # Try page title: "STW: SPDR S&P ASX 200 ETF | State Street ETFs"
            title = soup.title.string if soup.title else ''
            if code in title:
                name_match = re.match(rf'^{code}:\s*(.+?)\s*\|', title)
                if name_match:
                    name = name_match.group(1).strip()

        # SSGA pages embed data as HTML-entity-encoded JSON — decode once for all extractions
        decoded = html_module.unescape(resp.text)

        # MER from "total-expense-ratio" embedded JSON key
        mer = None
        mer_match = re.search(
            r'"total-expense-ratio"\s*:\s*\{[^}]*"value"\s*:\s*"([0-9.]+)\s*%',
            decoded,
        )
        if mer_match:
            mer = _safe_float(mer_match.group(1))
            if mer and not (0 < mer < 5):
                mer = None

        # Inception date
        inception = None
        inc_match = re.search(r'"inception-date"[^}]*"value"\s*:\s*"([^"]+)"', decoded)
        if inc_match:
            inception = inc_match.group(1).strip() or None

        # Benchmark from HTML table row: <td>Benchmark</td><td>{index name}</td>
        benchmark = None
        for td in soup.find_all('td'):
            if td.get_text(strip=True) == 'Benchmark':
                next_td = td.find_next_sibling('td')
                if next_td:
                    val = next_td.get_text(strip=True)
                    if val and len(val) > 3:
                        benchmark = val
                        break

        # Sector allocations from attrArray (picks the first occurrence)
        sectors = []
        sector_block = re.search(r'"attrArray"\s*:\s*(\[[^\]]+\])', decoded)
        if sector_block:
            try:
                attr_list = json.loads(sector_block.group(1))
                for entry in attr_list:
                    sector_name = (entry.get('name') or {}).get('value', '').strip()
                    weight_str  = (entry.get('weight') or {}).get('originalValue', '')
                    weight = _safe_float(weight_str)
                    if sector_name and weight is not None and 0 < weight < 100:
                        sectors.append({'sector': sector_name, 'weight_pct': weight})
            except (ValueError, json.JSONDecodeError, AttributeError):
                pass

        etf = {
            'code': code,
            'name': name,
            'issuer': 'SPDR',
            'expense_ratio': mer,
            'management_fee': mer,
            'inception_date': inception,
            'benchmark': benchmark,
            'exchange': 'ASX',
            'data_source': 'spdr',
            'issuer_url': url,
        }
        etf = {k: v for k, v in etf.items() if v is not None}
        upsert_etf(conn, etf)
        if sectors:
            upsert_sectors(conn, code, sectors, commit=False)

        # Holdings from SSGA daily Excel
        _scrape_ssga_holdings(conn, code)
        conn.commit()
        updated += 1

    duration = (datetime.utcnow() - started).total_seconds()
    log_scrape(conn, source, 'success' if updated else 'no_data',
               records_affected=updated, duration_secs=duration,
               started_at=started.isoformat())
    conn.close()
    logger.info(f"SPDR: updated {updated} ETFs")
    return updated


# ====================================================================
# StateStreet
# ====================================================================

# StateStreet ETFs listed on ASX/Cboe (not branded as SPDR)
_STATESTREET_AU_CODES = [
    'SPY', 'SSO', 'E200', 'SYI', 'QMIX', 'WXOZ', 'WXHG', 'OZR', 'OZF', 'WEMG',
]


def scrape_statestreet(db_path=None) -> int:
    """
    Download SSGA daily holdings Excel for each StateStreet AU ETF.
    Metadata (name, MER) comes from the ASX/Cboe reports; this adds holdings + sectors.
    """
    started = datetime.utcnow()
    source = 'statestreet'
    updated = 0

    conn = get_connection(db_path)

    # Fetch known StateStreet codes from DB (may be more than the hardcoded list)
    db_codes = [r[0] for r in conn.execute(
        "SELECT code FROM etfs WHERE issuer='StateStreet' ORDER BY code"
    ).fetchall()]
    codes = db_codes or _STATESTREET_AU_CODES

    for code in codes:
        if _scrape_ssga_holdings(conn, code):
            conn.commit()
            updated += 1
            logger.info(f"StateStreet: got holdings for {code}")
        else:
            logger.debug(f"StateStreet: no SSGA holdings for {code}")

    duration = (datetime.utcnow() - started).total_seconds()
    log_scrape(conn, source, 'success' if updated else 'no_data',
               records_affected=updated, duration_secs=duration,
               started_at=started.isoformat())
    conn.close()
    logger.info(f"StateStreet: updated {updated} ETFs")
    return updated


# ====================================================================
# Global X
# ====================================================================

_GLOBALX_STRAPI_BASE = 'https://d255kjlej81hb4.cloudfront.net'
_GLOBALX_FUNDS_PAGE  = 'https://www.globalxetfs.com.au/funds/ndq/'
_GLOBALX_CHUNK_RE    = re.compile(r'/_next/static/chunks/([^"\'?\s]+\.js)')
_GLOBALX_TOKEN_RE    = re.compile(
    r'Authorization.*?Bearer.*?concat\("([a-f0-9]{60,})', re.DOTALL
)

# Global X Strapi category name → canonical asset class
_GLOBALX_CATEGORY_MAP = {
    'core':              'Australian Equities',
    'income':            'Fixed Income',
    'international':     'International Equities',
    'thematic':          'Thematic',
    'commodities':       'Commodities',
    'crypto':            'Digital Assets',
    'leveraged and inverse': 'Alternatives',
    'leveraged':         'Alternatives',
}


def _globalx_get_token() -> str | None:
    """
    Discover the Global X Strapi bearer token from the Next.js JS bundles.
    The token is embedded in one of the static chunks (stable across page loads).
    """
    page = fetch(_GLOBALX_FUNDS_PAGE)
    if not page:
        return None
    chunks = _GLOBALX_CHUNK_RE.findall(page.text)
    for chunk in chunks:
        chunk_url = f'https://www.globalxetfs.com.au/_next/static/chunks/{chunk}'
        resp = fetch(chunk_url)
        if not resp:
            continue
        m = _GLOBALX_TOKEN_RE.search(resp.text)
        if m:
            return m.group(1)
    return None


def _parse_globalx_pcf(content: bytes, code: str) -> list[dict]:
    """
    Parse a Global X PCF (Portfolio Composition File) Excel.
    Structure:
      Header row: '#', 'Component Name', 'ISIN', 'SEDOL', 'Bloomberg Ticker',
                  'Number of Shares', 'Local CCY', 'Local CCY Price',
                  'Market Value (Base CCY)', 'Weight', 'Sector', 'Country'
    Weight is a decimal fraction (0.121 = 12.1%) — multiply by 100.
    """
    try:
        import openpyxl
    except ImportError:
        return []
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        logger.warning(f"Global X PCF parse error for {code}: {e}")
        return []

    ws = wb.active
    col = {}
    holdings = []
    for row in ws.iter_rows(values_only=True):
        vals = list(row)
        # Header row detection
        if not col:
            row_lower = [str(v or '').strip().lower() for v in vals]
            if 'component name' in row_lower or 'name' in row_lower:
                for i, h in enumerate(row_lower):
                    if h in ('component name', 'name'):
                        col.setdefault('name', i)
                    elif 'weight' in h and 'market' not in h:
                        col.setdefault('weight', i)
                    elif h == 'sector':
                        col.setdefault('sector', i)
                    elif h == 'country':
                        col.setdefault('country', i)
                    elif 'isin' in h:
                        col.setdefault('isin', i)
                    elif 'bloomberg' in h or 'ticker' in h:
                        col.setdefault('ticker', i)
            continue

        def gcol(key):
            idx = col.get(key)
            return vals[idx] if idx is not None and idx < len(vals) else None

        name = str(gcol('name') or '').strip()
        if not name or name.lower() in ('cash', '-', 'n/a', ''):
            continue
        weight_raw = _safe_float(gcol('weight'))
        if weight_raw is None or weight_raw <= 0:
            continue
        # Weight is stored as a decimal fraction (e.g. 0.121 = 12.1%)
        weight_pct = weight_raw * 100 if weight_raw < 1.5 else weight_raw

        holdings.append({
            'name':       name,
            'ticker':     str(gcol('ticker') or '').strip() or None,
            'isin':       str(gcol('isin') or '').strip() or None,
            'weight_pct': round(weight_pct, 4),
            'sector':     str(gcol('sector') or '').strip() or None,
            'country':    str(gcol('country') or '').strip() or None,
        })
    return holdings


def scrape_globalx(db_path=None) -> int:
    """
    Scrape Global X Australia via their Strapi CMS API + PCF holdings files.
    API:  GET https://d255kjlej81hb4.cloudfront.net/api/products?populate=*
    PCF:  each product.pcf is a direct xlsx download URL
    """
    started = datetime.utcnow()
    source = 'globalx'
    updated = 0

    conn = get_connection(db_path)

    # Step 1: Get auth token
    token = _globalx_get_token()
    if not token:
        logger.error("Global X: could not discover bearer token — skipping")
        log_scrape(conn, source, 'error', records_affected=0,
                   duration_secs=0, started_at=started.isoformat())
        conn.close()
        return 0

    auth_headers = {'Authorization': f'Bearer {token}'}

    # Step 2: Fetch all products from Strapi
    api_url = f'{_GLOBALX_STRAPI_BASE}/api/products?populate=*&pagination[limit]=100'
    resp = fetch(api_url, headers=auth_headers)
    if not resp:
        logger.error("Global X: Strapi API request failed")
        conn.close()
        return 0

    try:
        payload = resp.json()
        products = payload.get('data', [])
    except Exception as e:
        logger.error(f"Global X: Strapi JSON parse error: {e}")
        conn.close()
        return 0

    logger.info(f"Global X: {len(products)} products from Strapi API")

    for product in products:
        ticker = str(product.get('ticker') or '').strip().upper()
        if not ticker or not re.match(r'^[A-Z][A-Z0-9]{1,5}$', ticker):
            continue

        name = str(product.get('name') or '').strip() or None
        mer  = _safe_float(product.get('manageCosts'))
        inception = str(product.get('inceptionDate') or '').strip() or None
        slug = str(product.get('pageSlug') or ticker.lower())
        issuer_url = f'https://www.globalxetfs.com.au/funds/{slug}/'

        # Asset class from Strapi category
        cat_name = ((product.get('category') or {}).get('name') or '').strip().lower()
        asset_class = _GLOBALX_CATEGORY_MAP.get(cat_name) or normalise_asset_class(cat_name) or None

        etf = {
            'code':         ticker,
            'name':         name,
            'issuer':       'Global X',
            'expense_ratio': mer,
            'management_fee': mer,
            'asset_class':  asset_class,
            'inception_date': inception,
            'data_source':  'globalx',
            'issuer_url':   issuer_url,
            'exchange':     'ASX',
        }
        etf = {k: v for k, v in etf.items() if v is not None}
        upsert_etf(conn, etf)

        # Step 3: Download PCF for holdings
        pcf_url = str(product.get('pcf') or '').strip()
        if pcf_url:
            pcf_resp = fetch(pcf_url)
            if pcf_resp and pcf_resp.status_code == 200:
                rows = _parse_globalx_pcf(pcf_resp.content, ticker)
                if rows:
                    # Deduplicate by name
                    merged: dict[str, dict] = {}
                    for h in rows:
                        n = h['name']
                        if n in merged:
                            merged[n]['weight_pct'] = round(
                                (merged[n]['weight_pct'] or 0) + (h['weight_pct'] or 0), 4
                            )
                        else:
                            merged[n] = h
                    upsert_holdings(conn, ticker, list(merged.values()), commit=False)

                    # Sector aggregation
                    sector_wt: dict[str, float] = {}
                    for h in merged.values():
                        s = h.get('sector')
                        if s and s.lower() not in ('cash', '-', 'n/a', ''):
                            sector_wt[s] = round(sector_wt.get(s, 0) + (h.get('weight_pct') or 0), 4)
                    if sector_wt:
                        upsert_sectors(conn, ticker,
                                       [{'sector': s, 'weight_pct': w} for s, w in sector_wt.items()],
                                       commit=False)

        conn.commit()
        updated += 1

    duration = (datetime.utcnow() - started).total_seconds()
    log_scrape(conn, source, 'success' if updated else 'no_data',
               records_affected=updated, duration_secs=duration,
               started_at=started.isoformat())
    conn.close()
    logger.info(f"Global X: updated {updated} ETFs")
    return updated


# ====================================================================
# CXA Issuer Scrapers (distribution yield)
# ====================================================================

def _extract_distribution_yield(soup, raw_text: str) -> float | None:
    """
    Try to extract a distribution yield percentage from a fund page.
    Checks JSON-LD structured data, meta tags, HTML tables, then raw-text regex.
    Returns the yield as a float (e.g. 4.5 for 4.5%), or None if not found.
    """
    # JSON-LD structured data
    for script in soup.find_all('script', type='application/ld+json'):
        try:
            data = json.loads(script.string.strip())
            for key in ('distributionYield', 'distribution_yield', 'dividendYield', 'yield'):
                val = _safe_float(data.get(key))
                if val and 0 < val < 50:
                    return val
        except (ValueError, json.JSONDecodeError, AttributeError):
            continue

    # Meta tags
    for meta in soup.find_all('meta'):
        name_attr = (meta.get('name') or meta.get('property') or '').lower()
        if 'yield' in name_attr or 'distribution' in name_attr:
            val = _safe_float(meta.get('content'))
            if val and 0 < val < 50:
                return val

    # HTML tables with distribution/yield headers
    for table in soup.find_all('table'):
        headers = [th.get_text(strip=True).lower() for th in table.find_all('th')]
        if not any('distribution' in h or 'yield' in h for h in headers):
            continue
        for row in table.find_all('tr'):
            cells = [td.get_text(strip=True) for td in row.find_all(['td', 'th'])]
            if len(cells) < 2:
                continue
            label = cells[0].lower()
            if 'distribution yield' in label or ('yield' in label and 'distribution' in label):
                val = _safe_float(cells[1])
                if val and 0 < val < 50:
                    return val

    # Raw text regex fallback — ordered most-to-least specific
    for pattern in [
        # Coolabah "Currently yielding 6.2% p.a." callout
        r'currently\s+yielding\s+([0-9]+\.?[0-9]*)\s*%',
        # Generic distribution yield label
        r'distribution\s+yield[^<>0-9]*([0-9]+\.?[0-9]*)\s*%',
        r'"distributionYield"\s*:\s*"?([0-9]+\.?[0-9]*)',
        # Running yield / gross yield table rows
        r'(?:running|gross)\s+(?:running\s+)?yield[^<>0-9]*([0-9]+\.?[0-9]*)\s*%',
    ]:
        m = re.search(pattern, raw_text, re.I)
        if m:
            val = _safe_float(m.group(1))
            if val and 0 < val < 50:
                return val

    return None


def _make_slug(name: str) -> str:
    """Convert a fund name to a URL-friendly slug."""
    slug = re.sub(r'[^a-z0-9\s]', '', name.lower())
    return re.sub(r'\s+', '-', slug).strip('-')


# Hardcoded URL slugs for CXA issuers where auto-generation from name doesn't match
# the actual site structure.

# Coolabah: domain is coolabahcapital.com (not coolabah.com.au); slugs are non-standard
_COOLABAH_SLUGS = {
    'CBNX': 'coolabah-global-carbon-leaders-complex-etf',
    'FIXD': 'active-composite-bond-strategy',
    'FRNS': 'coolabah-short-term-income-fund-managed-fund',
    'YLDX': 'coolabah-global-floating-rate-high-yield-fund',
}

# JPMorgan: fund detail URLs require the ISIN appended to the name slug
_JPMORGAN_ISINS = {
    'JPGB': 'au0000302440',
    'JPIE': 'au0000282345',
}

# Schroders: URL uses the ASX code (lowercase) under a per-fund category segment
_SCHRODERS_CATEGORIES = {
    'HIGH': 'active-etf',
    'PAYS': 'fixed-income',
}


def scrape_cxa_issuers(db_path=None) -> int:
    """
    Enrich CXA ETFs with distribution_yield scraped from issuer websites.
    Targets Coolabah, PIMCO, JPMorgan, Janus Henderson, and Schroders.
    PIMCO/JPMorgan/Janus Henderson/Schroders load fund data dynamically so
    will typically return no yield; Coolabah embeds yield in static HTML.
    Logs failures without crashing.
    """
    started = datetime.utcnow()
    source = 'cxa_issuers'
    updated = 0

    try:
        from bs4 import BeautifulSoup
    except ImportError:
        logger.warning("beautifulsoup4 required for CXA issuer scrapers")
        return 0

    conn = get_connection(db_path)

    rows = conn.execute(
        "SELECT code, name, issuer FROM etfs WHERE exchange = 'CXA' AND issuer IS NOT NULL"
    ).fetchall()

    if not rows:
        logger.info("CXA issuers: no CXA ETFs in DB")
        conn.close()
        return 0

    # Group by issuer
    by_issuer: dict[str, list[dict]] = {}
    for row in rows:
        issuer = row['issuer']
        if issuer not in by_issuer:
            by_issuer[issuer] = []
        by_issuer[issuer].append({'code': row['code'], 'name': row['name'] or ''})

    # --- Per-issuer URL builders ---

    def _coolabah_url(code, name):
        slug = _COOLABAH_SLUGS.get(code) or _make_slug(name)
        return f"https://coolabahcapital.com/{slug}/"

    def _pimco_url(code, name):
        # Replace "Active ETF" suffix with "Fund" (keep "PIMCO " prefix in slug)
        slug_name = re.sub(r'\s+active\s+etf\s*$', ' Fund', name, flags=re.I)
        slug = _make_slug(slug_name)
        return f"https://www.pimco.com/au/en/investments/etf/{slug}/ausetf-aud"

    def _jpmorgan_url(code, name):
        isin = _JPMORGAN_ISINS.get(code)
        if not isin:
            return None
        slug = _make_slug(name)
        return f"https://am.jpmorgan.com/au/en/asset-management/adv/products/{slug}-{isin}"

    def _janus_url(code, name):
        # Strip "Janus Henderson " prefix from the name slug
        slug_name = re.sub(r'^janus\s+henderson\s+', '', name, flags=re.I)
        slug = _make_slug(slug_name)
        return f"https://www.janushenderson.com/en-au/adviser/{slug}/"

    def _schroders_url(code, name):
        category = _SCHRODERS_CATEGORIES.get(code, 'active-etf')
        return f"https://www.schroders.com/en-au/au/adviser/funds/{category}/{code.lower()}/"

    issuer_configs = [
        ('Coolabah', _coolabah_url),
        ('PIMCO', _pimco_url),
        ('JPMorgan', _jpmorgan_url),
        ('Janus Henderson', _janus_url),
        ('Schroders', _schroders_url),
    ]

    for issuer_name, url_fn in issuer_configs:
        funds = by_issuer.get(issuer_name, [])
        if not funds:
            continue

        logger.info(f"CXA issuers: scraping {issuer_name} ({len(funds)} funds)")
        for fund in funds:
            code = fund['code']
            name = fund['name']
            try:
                url = url_fn(code, name)
                if not url:
                    logger.debug(f"  {code}: no URL available, skipping")
                    continue

                resp = fetch(url)
                if not resp or resp.status_code >= 400:
                    logger.debug(f"  {code}: no usable response from {url}")
                    continue

                soup = BeautifulSoup(resp.text, 'html.parser')
                dist_yield = _extract_distribution_yield(soup, resp.text)

                if dist_yield is not None:
                    upsert_etf(conn, {
                        'code': code,
                        'distribution_yield': dist_yield,
                        'data_source': f'cxa_{issuer_name.lower().replace(" ", "_")}',
                    })
                    updated += 1
                    logger.info(f"  {code}: distribution_yield={dist_yield}")
                else:
                    logger.debug(f"  {code}: distribution yield not found at {url}")

            except Exception as e:
                logger.warning(f"CXA issuers: error scraping {issuer_name} {code}: {e}")

        conn.commit()

    duration = (datetime.utcnow() - started).total_seconds()
    log_scrape(conn, source, 'success' if updated else 'no_data',
               records_affected=updated, duration_secs=duration,
               started_at=started.isoformat())
    conn.close()
    logger.info(f"CXA issuers: updated {updated} ETFs with distribution yield")
    return updated


# ====================================================================
# Orchestrator
# ====================================================================

def scrape_all_issuers(db_path=None) -> int:
    """Run all issuer scrapers. Returns total ETFs updated."""
    total = 0
    scrapers = [
        ('BetaShares', scrape_betashares),
        ('VanEck', scrape_vaneck),
        ('Vanguard', scrape_vanguard),
        ('iShares', scrape_ishares),
        ('SPDR', scrape_spdr),
        ('StateStreet', scrape_statestreet),
        ('Global X', scrape_globalx),
        ('CXA Issuers', scrape_cxa_issuers),
    ]

    for name, func in scrapers:
        try:
            count = func(db_path)
            total += count
        except Exception as e:
            logger.error(f"Issuer scraper '{name}' failed: {e}", exc_info=True)

    # Update issuer stats after all scrapers run
    conn = get_connection(db_path)
    update_issuer_stats(conn)
    conn.close()

    logger.info(f"All issuer scrapers complete: {total} total ETFs updated")
    return total
